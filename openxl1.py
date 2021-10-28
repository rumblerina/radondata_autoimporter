from openpyxl import Workbook, load_workbook
from dbfread import DBF

wb = Workbook()
table = DBF("DBRADON.DBF", load = True, ignore_missing_memofile=True)
length = len(table)
ws1 = wb.create_sheet("RFD", 0)
ws2 = wb.create_sheet("RAC")
ws1['A1'] = "Date"
ws1['E1'] = "Measurement Start"
ws1['F1'] = "Measurement End"
ws1['G1'] = "Exposition Length"
ws1['H1'] = "Measurement Length"
ws1['I1'] = "Radon activity"
ws1['J1'] = "Radon activity +-"
ws1['K1'] = "RFD"
ws1['L1'] = "RFD+-"
ws1['O1'] = "SK-13"

ws2['A1'] = "Date"
ws2['E1'] = "Measurement Length"
ws2['F1'] = "Radon activity"
ws2['G1'] = "Radon activity +-"
ws2['H1'] = "RAC"
ws2['I1'] = "RAC+-"
ws2['J1'] = "Volume, l"
ws2['L1'] = "SK-13"

# Imports RFD measurement time data************************************
c = 1 # counter for loop
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        timeb = str(table.records[i]['DTEXP'])
        ws1['A' + str(c)].value = timeb
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        timeb = str(table.records[i]['BGEXP']) + ' ' + str(table.records[i]['BGTM']) + ':' + str(table.records[i]['BGTM2'])
        ws1['E' + str(c)].value = timeb
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        timee = str(table.records[i]['EDEXP']) + ' ' + str(table.records[i]['EDTM']) + ':' + str(table.records[i]['EDTM2'])
        ws1['F' + str(c)].value = timee           
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['G' + str(c)].value = table.records[i]['TEXP']
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['H' + str(c)].value = table.records[i]['TL']            
print("Importation of RFD time data complete")

# Imports RFD radon activity************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['I' + str(c)].value = table.records[i]['A214BI']
# Imports radon activity measurement error
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['J' + str(c)].value = table.records[i]['D214BI']            
print("Importation of RFD radon activity data complete")

# Imports rfd**************************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['K' + str(c)].value = table.records[i]['A_RN1']
# Imports rfd measurement error
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['L' + str(c)].value = table.records[i]['D_RN1']            
print("Importation of radon flux data complete")

# Imports RFD SK13 values***********************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 30:
        c = c + 1
        ws1['O' + str(c)].value = table.records[i]["MEASURE"]      
print("Importation of RFD SK13 data complete")

# Imports RAC time values*********************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        timeb = str(table.records[i]['DTEXP'])
        ws2['A' + str(c)].value = timeb  
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['E' + str(c)].value = table.records[i]['TL']     
print("Importation of RAC time data complete")

# Imports RAC radon charcoal activity************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['F' + str(c)].value = table.records[i]['A214BI']
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['G' + str(c)].value = table.records[i]['D214BI']            
print("Importation of RFD radon activity data complete")

# Imports RAC values**************************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['H' + str(c)].value = table.records[i]['A_RN1']
# Imports RAC measurement error
c = 1
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['I' + str(c)].value = table.records[i]['D_RN1']            
c = 1
# Imports RAC measurement volume
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['J' + str(c)].value = table.records[i]['VPR']            
print("Importation of RAC data complete")
# Imports RAC SK13 values***********************************************
c = 1 
for i in range(length):
    if table.records[i]['TYPE'] == 60:
        c = c + 1
        ws2['L' + str(c)].value = table.records[i]["MEASURE"]      
print("Importation of RAC SK13 data complete")

wb.save("radon autodata.xlsx")
